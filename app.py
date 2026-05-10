"""
Panamerican Coffee Trade - Reportes Dinámicos
Versión producción v1.0
"""

from __future__ import annotations

import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from eximware_parser import process_all


def _write_table_sheet(
    df: pd.DataFrame,
    writer: pd.ExcelWriter,
    sheet_name: str,
    table_name: str,
) -> None:
    """Escribe el DataFrame en una hoja con formato de tabla nativo de Excel
    (equivalente a Ctrl+T): filtros, banding y rango con nombre.
    """
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    if df.empty:
        return

    n_rows, n_cols = df.shape
    last_col_letter = get_column_letter(n_cols)
    ref = f"A1:{last_col_letter}{n_rows + 1}"

    safe_name = re.sub(r"[^A-Za-z0-9_]", "", table_name) or "Reporte"
    if safe_name[0].isdigit():
        safe_name = "T_" + safe_name

    table = Table(displayName=safe_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(table)

    for i, col in enumerate(df.columns, start=1):
        try:
            max_data_len = (
                df[col].astype(str).map(len).max() if len(df) else 0
            )
        except Exception:
            max_data_len = 0
        col_width = min(max(int(max_data_len), len(str(col))) + 2, 40)
        ws.column_dimensions[get_column_letter(i)].width = max(col_width, 10)

st.set_page_config(
    page_title="Panamerican · Reportes Dinámicos",
    page_icon="☕",
    layout="wide",
)

# ---- Auth simple ------------------------------------------------------------

def _check_password() -> bool:
    """Bloquea el acceso hasta que el usuario ingrese la contraseña correcta."""
    if st.session_state.get("autenticado"):
        return True

    def _validar():
        if st.session_state["_pwd_input"] == st.secrets.get("password", ""):
            st.session_state["autenticado"] = True
        else:
            st.session_state["pwd_incorrecto"] = True

    with st.container():
        st.markdown("## ☕ Panamerican · Reportes Dinámicos")
        st.markdown("---")
        st.text_input(
            "Contraseña de acceso",
            type="password",
            key="_pwd_input",
            on_change=_validar,
        )
        if st.session_state.get("pwd_incorrecto"):
            st.error("Contraseña incorrecta. Intentá de nuevo.")
    return False


if not _check_password():
    st.stop()

# ---- Estilos ----------------------------------------------------------------
st.markdown(
    """
    <style>
      .main .block-container { padding-top: 1.2rem; padding-bottom: 2rem; }
      h1 { margin-bottom: 0.2rem; }
      .version-tag { color: #888; font-size: 0.85rem; }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("☕ Panamerican · Reportes Dinámicos")
st.markdown(
    '<div class="version-tag">v1.0 · Cargá los reportes crudos de Eximware y explorá al vuelo — sin limpiar Excel.</div>',
    unsafe_allow_html=True,
)

# ---- Sidebar: carga de archivos ---------------------------------------------
with st.sidebar:
    st.header("1. Cargar reportes de Eximware")
    st.caption(
        "Descargá los reportes directamente de Eximware y subílos acá. "
        "No hay que limpiarlos — el motor lo hace automáticamente."
    )

    alloc_llc = st.file_uploader("Allocation · LLC", type=["xlsx"], key="al_llc")
    alloc_pct = st.file_uploader("Allocation · PCT", type=["xlsx"], key="al_pct")
    alloc_3rd = st.file_uploader("Allocation · 3ra empresa (opcional)", type=["xlsx"], key="al_3rd")

    st.markdown("---")

    pos_llc = st.file_uploader("Position · LLC", type=["xlsx"], key="po_llc")
    pos_pct = st.file_uploader("Position · PCT", type=["xlsx"], key="po_pct")
    pos_3rd = st.file_uploader("Position · 3ra empresa (opcional)", type=["xlsx"], key="po_3rd")

    st.markdown("---")
    run = st.button("▶ Procesar reportes", type="primary", use_container_width=True)


# ---- Pipeline con cache -----------------------------------------------------

@st.cache_data(show_spinner="Procesando archivos de Eximware...")
def _run_pipeline(alloc_sources: list[tuple[bytes, str]], pos_sources: list[tuple[bytes, str]]):
    alloc_files = [(io.BytesIO(b), c) for b, c in alloc_sources]
    pos_files   = [(io.BytesIO(b), c) for b, c in pos_sources]
    return process_all(alloc_files, pos_files)


def _read_upload(up, label):
    return (up.getvalue(), label) if up is not None else None


# ---- Estado -----------------------------------------------------------------

if "result" not in st.session_state:
    st.session_state.result = None

if run:
    alloc_sources = [x for x in [
        _read_upload(alloc_llc, "LLC"),
        _read_upload(alloc_pct, "PCT"),
        _read_upload(alloc_3rd, "3RA"),
    ] if x is not None]
    pos_sources = [x for x in [
        _read_upload(pos_llc, "LLC"),
        _read_upload(pos_pct, "PCT"),
        _read_upload(pos_3rd, "3RA"),
    ] if x is not None]

    if not alloc_sources or not pos_sources:
        st.error("Necesito al menos 1 archivo de Allocation y 1 de Position para procesar.")
    else:
        st.session_state.result = _run_pipeline(tuple(alloc_sources), tuple(pos_sources))

if st.session_state.result is None:
    st.info("⬅️ Subí los archivos de Eximware en el panel izquierdo y presioná **Procesar reportes**.")
    st.stop()


# ---- Datos procesados -------------------------------------------------------

df: pd.DataFrame = st.session_state.result["reporte_final"].copy()

df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce")
for dt_col in ("BL Date", "Period start", "Period end", "Contract Dt.", "Input Dt."):
    if dt_col in df.columns:
        df[dt_col] = pd.to_datetime(df[dt_col], errors="coerce")
# ETD se mantiene como texto (Eximware lo exporta como "MM/DD/YYYY ETD" / "MM/DD/YYYY ETA")
# No convertir a datetime o se pierde toda la información.

# ---- Filtros dinámicos ------------------------------------------------------
st.subheader("1. Filtros")

filter_cols = [
    "Company", "P/S", "Counter Party", "Alloc. Party",
    "Origin", "Product Class", "Price Status", "Account Representative",
]

with st.expander("Filtros por categoría", expanded=True):
    cols = st.columns(4)
    filter_values: dict[str, list] = {}
    for i, col in enumerate(filter_cols):
        if col not in df.columns:
            continue
        options = sorted(df[col].dropna().astype(str).unique())
        with cols[i % 4]:
            sel = st.multiselect(col, options, key=f"f_{col}")
            if sel:
                filter_values[col] = sel

with st.expander("Filtros de fecha (BL Date)", expanded=False):
    c1, c2, c3 = st.columns(3)
    bl_mode = c1.radio(
        "BL Date",
        ["Cualquiera", "Con fecha (embarcado)", "Sin fecha (pendiente)", "Rango"],
        horizontal=False,
        key="bl_mode",
    )
    bl_from = c2.date_input("Desde", value=None, key="bl_from")
    bl_to   = c3.date_input("Hasta", value=None, key="bl_to")

# Aplicar filtros
filtered = df.copy()
for col, vals in filter_values.items():
    filtered = filtered[filtered[col].astype(str).isin(vals)]

if "BL Date" in filtered.columns:
    if bl_mode == "Con fecha (embarcado)":
        filtered = filtered[filtered["BL Date"].notna()]
    elif bl_mode == "Sin fecha (pendiente)":
        filtered = filtered[filtered["BL Date"].isna()]
    elif bl_mode == "Rango":
        if bl_from:
            filtered = filtered[filtered["BL Date"] >= pd.Timestamp(bl_from)]
        if bl_to:
            filtered = filtered[filtered["BL Date"] <= pd.Timestamp(bl_to)]

# ---- KPIs -------------------------------------------------------------------
st.markdown("---")
st.subheader("2. Resumen")

k1, k2, k3, k4 = st.columns(4)
total_rows = len(df)
total_qty  = df["Quantity"].sum()
filt_rows  = len(filtered)
filt_qty   = filtered["Quantity"].sum()

k1.metric("Registros",      f"{filt_rows:,}",    delta=f"de {total_rows:,} totales",  delta_color="off")
k2.metric("Purchases (P)",  f"{(filtered['P/S'] == 'P').sum():,}")
k3.metric("Sales (S)",      f"{(filtered['P/S'] == 'S').sum():,}")
k4.metric("Quantity",       f"{filt_qty:,.0f}",  delta=f"de {total_qty:,.0f} totales", delta_color="off")

# ---- Tabla ------------------------------------------------------------------
st.subheader(f"3. Resultado · {len(filtered):,} filas")

default_cols = [
    "Company", "Ref #", "P/S", "Counter Party", "Origin", "Grade",
    "Product Class", "Quantity", "UOM", "BL Date", "ETD", "Alloc. Party",
    "Allocation", "Price Status", "Final Price",
]
available      = [c for c in df.columns if c not in {"__index_level_0__"}]
default_display = [c for c in default_cols if c in available]

with st.expander("Elegí columnas a mostrar", expanded=False):
    show_cols = st.multiselect(
        "Columnas visibles",
        options=available,
        default=default_display,
        key="show_cols",
    )

if not show_cols:
    show_cols = default_display

st.dataframe(filtered[show_cols], use_container_width=True, height=420)

# ---- Gráfico ----------------------------------------------------------------
st.subheader("4. Gráfico de cantidades")

g1, g2 = st.columns(2)
group_col = g1.selectbox(
    "Agrupar por",
    options=["Origin", "Counter Party", "Alloc. Party", "Grade", "Product Class", "Price Status", "Company"],
    index=0,
    key="grp",
)
top_n = g2.slider("Top N", 5, 30, 10, key="topn")

if len(filtered) > 0 and group_col in filtered.columns:
    chart_df = (
        filtered.groupby(group_col, dropna=False)["Quantity"]
        .sum()
        .sort_values(ascending=False)
        .head(top_n)
    )
    st.bar_chart(chart_df)
else:
    st.warning("Sin datos para graficar con los filtros actuales.")

# ---- Descarga ---------------------------------------------------------------
st.subheader("5. Exportar")
st.caption(
    f"Filtrado: **{len(filtered):,} filas** con columnas visibles · "
    f"Total: **{len(df):,} filas** con todas las columnas."
)

_stamp = datetime.now().strftime("%Y%m%d_%H%M")

col_filt, col_all = st.columns(2)

buf_filt = io.BytesIO()
with pd.ExcelWriter(buf_filt, engine="openpyxl") as writer:
    _write_table_sheet(
        filtered[show_cols],
        writer,
        sheet_name="Reporte filtrado",
        table_name="ReporteFiltrado",
    )
col_filt.download_button(
    f"📥 Descargar FILTRADO ({len(filtered):,} filas)",
    data=buf_filt.getvalue(),
    file_name=f"panamerican_filtrado_{_stamp}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

buf_all = io.BytesIO()
with pd.ExcelWriter(buf_all, engine="openpyxl") as writer:
    _write_table_sheet(
        df,
        writer,
        sheet_name="Reporte completo",
        table_name="ReporteCompleto",
    )
col_all.download_button(
    f"📦 Descargar TOTAL ({len(df):,} filas, todas las columnas)",
    data=buf_all.getvalue(),
    file_name=f"panamerican_total_{_stamp}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
